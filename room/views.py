from django.shortcuts import render,redirect
from .models import Room,Quiz,Score
import string
import random
from django.core.cache import cache
from django.http.response import JsonResponse
import json
from .participant_middleware import get_user_middleware
from pathlib import Path
from time import sleep

# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = Path(__file__).resolve().parent.parent
from openpyxl import load_workbook
# Create your views here.

# ye view testing purpose ke liye h
def quiz_page(request,group_name):
    room, created = Room.objects.get_or_create(name=group_name)
    if created:
        print('New Group Created')
        # cache_group_name='user_count_'+group_name
        # cache.set(cache_group_name, 0,246060) # 24 hrs TTL (in seconds)
        quiz=Quiz.objects.create(host=request.user,room=room)
    else:
        print('Existing Group')
        # cache_group_name='user_count_'+group_name
        # cache.set(cache_group_name, 0,246060) # 24 hrs TTL (in seconds)
        print('user_count_'+str(cache.get('user_count_'+group_name)))
    quiz=Quiz.objects.get(room=room)
    # print(quiz.questions.all())
    questions=[]
    print(questions)
    return render(request,'app/manual_quiz_page.html',{'group_name':group_name,'questions':quiz.questions.all()})

def quiz(request):
    group_name=None
    while True:
        group_name = ''.join(random.choices(string.ascii_uppercase+string.digits, k=6))
        group = Room.objects.filter(name=group_name).exists()
        if not group:
            break
    cache_group_name='user_count_'+group_name
    cache.set(cache_group_name, 0,246060) # 24 hrs TTL (in seconds)
    return redirect('quiz_page', group_name=group_name)

def auto_quiz(request):
    group_name=None
    while True:
        group_name = ''.join(random.choices(string.ascii_uppercase+string.digits, k=6))
        group = Room.objects.filter(name=group_name).exists()
        if not group:
            break
    cache_group_name='user_count_'+group_name
    cache.set(cache_group_name, 0,246060) # 24 hrs TTL (in seconds)
    return redirect('auto_quiz_page', group_name=group_name)

def auto_quiz_page(request,group_name):
    room, created = Room.objects.get_or_create(name=group_name)
    if created:
        print('New Group Created')
        # cache_group_name='user_count_'+group_name
        # cache.set(cache_group_name, 0,246060) # 24 hrs TTL (in seconds)
        quiz=Quiz.objects.create(host=request.user,room=room)
    else:
        print('Existing Group')
        # cache_group_name='user_count_'+group_name
        # cache.set(cache_group_name, 0,246060) # 24 hrs TTL (in seconds)
        print('user_count_'+str(cache.get('user_count_'+group_name)))
    
    questions=get_excel_data(request)
    print(questions)
    return render(request,'app/auto_quiz_page.html',{'group_name':group_name,'questions':questions})

def get_excel_data(self):
    try:
        '''
        {
        question: Q,
        answers: [
            { "option": op1, "correct": cop1 },
            { "option": op2, "correct": cop2 },
            { "option": op3, "correct": cop3 },
            { "option": op4, "correct": cop4 },
        ]
        }
        '''
        Question=[]
        myfile=load_workbook(BASE_DIR/"static/quiz.xlsx")
        s=myfile['Sheet1']
        row=s.max_row
        column=s.max_column
        for r in range(2,row-1):
            mcq=s.cell(row=r,column=1).value
            for c in range(1,5):
                ans=s.cell(row=r,column=5).value
            Question.append({"question":mcq,"answers":ans})       
        pass
    except Exception as e:
        print(e)
    return Question

def home(request):
    print('sdfdd',request.user.is_authenticated)
    # create room and join room wala page
    return render(request,'app/home.html')

def join_room(request):
    json_data = json.loads(request.body.decode("utf-8"))
    try:
        quiz=Room.objects.filter(name=json_data["room_id"])[0]
        print(quiz)
        return JsonResponse(data={"exist":True},safe=False)
    except:
        print('room not exists')
        return JsonResponse(data={"exist":False},safe=False)

@get_user_middleware
def submit_quiz_answers(request):
    if request.method == 'POST':
        try:
            print(request.user.first_name)
            print(json.loads(request.body.decode("utf-8")))
            json_data = json.loads(request.body.decode("utf-8"))
            roomid=json_data['roomid']
            points=json_data['points']
            room=Room.objects.filter(name=roomid)[0]
            print(room)
            Score.objects.create(room=room,user=request.user,point=points)
            sleep(2)
            score_obj=Score.objects.filter(room=room).values('user__first_name','user__last_name','point').order_by('-point')
            print(score_obj)
            result=[]
            for obj in score_obj:
                t={
                    'user_name':str(obj['user__first_name']+' '+obj['user__last_name']),
                    'point':obj['point']
                }
                result.append(t)
            return JsonResponse(data={'status':True,'message':'Successfully Submit','data':result})
        except Exception as e:
            print(e)
            return JsonResponse(data={'status':False,'message':'Failed to submit'})
    else:
        return JsonResponse(data={'status':False,'message':'Failed to submit'})
    