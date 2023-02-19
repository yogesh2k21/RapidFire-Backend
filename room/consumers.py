from channels.generic.websocket import AsyncWebsocketConsumer
from .models import Room, Chat,Option,Quiz,MCQ
import json
from channels.db import database_sync_to_async
from django.core.cache import cache
from pathlib import Path
import random
from openpyxl import load_workbook
BASE_DIR = Path(__file__).resolve().parent.parent


def online_user_count(group_name,action):
    if action == 'INCR':
        cache.incr('user_count_'+group_name)
    elif action == 'DECR':
        cache.decr('user_count_'+group_name)
    return cache.get('user_count_'+group_name)

def is_quiz_ending(room):
        print('inside end quiz fucntion')
        q = Quiz.objects.filter(room=room)[0]
        q.completed=True
        q.save()
        cache.delete('user_count_'+str(room.name))
        print('room user count cleared from cache')
        print('quiz ended successfully')

# In this Class we are reading Question and Options from excel file
class AutoAsyncConsumer(AsyncWebsocketConsumer):

    def get_mcq(self,mcq_id,room):
        myfile=load_workbook(BASE_DIR/"static/quiz.xlsx")
        s=myfile['Sheet1']

        question=s.cell(row=mcq_id,column=1).value
        mcq=MCQ(problem_statement=question)
        new_op=Option.objects.create(statement=s.cell(row=mcq_id,column=5).value,valid=True)
        new_op.save()
        # mcq.options.set(new_op)
        mcq.correct=new_op
        mcq.save()
        quiz=Quiz.objects.filter(room=room)[0]
        quiz.questions.add(mcq)
        quiz.save()
        answers=[]
        for c in range(2,6):
            if c == 5:
                t={"option":s.cell(row=mcq_id,column=c).value,"correct":True}       
            else:
                t={"option":s.cell(row=mcq_id,column=c).value,"correct":False}
            answers.append(t)
        random.shuffle(answers)
        data=dict({"question":question,"answers":answers})    
        print(data)
        return data

    async def connect(self):
        print("web socket connected....")
        print("channel_name " + self.channel_name)
        print("channel_layer ", self.channel_layer)
        self.group_name = self.scope["url_route"]["kwargs"]["group_name"]
        # print('group_name '+self.scope['url_route']['kwargs']['group_name'])
        print("group_name ", self.group_name)
        # adding this channel to this given group_name which is come within the url as parameter
        await self.channel_layer.group_add(self.group_name, self.channel_name)
        online_users=None
        if int(cache.get('user_count_'+str(self.group_name)))<5:
            online_users=online_user_count(str(self.group_name),'INCR')
        elif int(cache.get('user_count_'+str(self.group_name)))==5:
            print('5 user connected')
            raise Exception("Sorry, Group is housefull now!!!")
        # sending online user count when someone join the group/room
        response = {"type": "chat.message", "message": "Someone joined","count":online_users,"endQuiz":False}
        await self.channel_layer.group_send(self.group_name, response)
        await self.accept()

    async def receive(self, text_data=None, bytes_data=None):
        print("Message received from client...", text_data,type(text_data))
        data = json.loads(text_data)
        print(data)

        status=data["status"]
        try:
            message = data["message"]
        except:
            message=""
            
        print(self.scope["user"])
        try:
            room = await database_sync_to_async(Room.objects.get)(name=self.group_name)
            isQuizEnd=data["endQuiz"]
            if isQuizEnd:
                await database_sync_to_async(is_quiz_ending)(room)
                response = {"type": "chat.message", "message": message,"count":cache.get('user_count_'+self.group_name),"endQuiz":True}
                await self.channel_layer.group_send(self.group_name, response)
                await self.disconnect(code=1001)

            # quiz_data=json.loads(data["message"]) 
            mcq_no=data["message"]
            message=await database_sync_to_async(self.get_mcq)(mcq_no+1,room)
            #broadcasting
            response = {"type": "chat.message", "message": message,"count":cache.get('user_count_'+self.group_name),"endQuiz":False}
            await self.channel_layer.group_send(self.group_name, response)
        except Exception as e:
            print('error found',e)
        #     await self.send(text_data=json.dumps({"message": "Login Required"}))

    async def chat_message(self, event):
        print("Event..", event)
        await self.send(text_data=json.dumps({"message": event["message"],"count":event["count"],"endQuiz":event["endQuiz"]}))

    async def disconnect(self, code):
        print("web socket disconnected ", code)
        print("channel_name " + self.channel_name)
        print("channel_layer ", self.channel_layer)
        try:
            online_users=online_user_count(self.group_name,'DECR')
        except:
            online_users=0
        await self.channel_layer.group_discard(self.group_name, self.channel_name)
        # sending online user count when someone left the group/room
        response = {"type": "chat.message", "message": "Someone Left","count":online_users,"endQuiz":False}
        await self.channel_layer.group_send(self.group_name, response)

# In this Class we are taking Question and Options from host portal
class CustomAsyncConsumer(AsyncWebsocketConsumer):

    def save_mcq_options(self,quiz_data,room):
        question=quiz_data["question"]
        options=quiz_data["answers"]
        mcq=MCQ(problem_statement=question)
        option_objects_list=[]
        for o in options:
            new_op=Option.objects.filter(statement=o['option'],valid=o['correct'])
            if new_op.count()==0:
                new_op=Option(statement=o['option'],valid=o['correct'])
                new_op.save()
            else:
                new_op=new_op[0]
                
            option_objects_list.append(new_op)
            if o['correct']:
                mcq.correct=new_op
        mcq.save()
        mcq.options.set(option_objects_list)
        # mcq.save()
        quiz=Quiz.objects.filter(room=room)[0]
        quiz.questions.add(mcq)
        quiz.save()
        print('mcq saved and return successfully')
        return mcq

    async def connect(self):
        print("web socket connected....")
        print("channel_name " + self.channel_name)
        print("channel_layer ", self.channel_layer)
        self.group_name = self.scope["url_route"]["kwargs"]["group_name"]
        # print('group_name '+self.scope['url_route']['kwargs']['group_name'])
        print("group_name ", self.group_name)
        # adding this channel to this given group_name which is come within the url as parameter
        await self.channel_layer.group_add(self.group_name, self.channel_name)
        online_users=None
        if int(cache.get('user_count_'+str(self.group_name)))<5:
            online_users=online_user_count(str(self.group_name),'INCR')
        elif int(cache.get('user_count_'+str(self.group_name)))==5:
            print('5 user connected')
            raise Exception("Sorry, Group is housefull now!!!")
        # sending online user count when someone join the group/room
        response = {"type": "chat.message", "message": "Someone joined","count":online_users,"endQuiz":False}
        await self.channel_layer.group_send(self.group_name, response)
        await self.accept()
        # await self.send(text_data=json.dumps({"message": "Someone joined","count":self.user_count}))

    async def receive(self, text_data=None, bytes_data=None):
        print("Message received from client...", text_data,type(text_data))
        data = json.loads(text_data)
        print(data)

        status=data["status"]
        try:
            message = data["message"]
        except:
            message=""
            
        print(self.scope["user"])
        try:
            room = await database_sync_to_async(Room.objects.get)(name=self.group_name)
            isQuizEnd=data["endQuiz"]
            if isQuizEnd:
                await database_sync_to_async(is_quiz_ending)(room)
                response = {"type": "chat.message", "message": message,"count":cache.get('user_count_'+self.group_name),"endQuiz":True}
                await self.channel_layer.group_send(self.group_name, response)
                await self.disconnect(code=1001)
                
            quiz_data=json.loads(data["message"]) 
            mcq=await database_sync_to_async(self.save_mcq_options)(quiz_data,room)
            chat = Chat(content=message, room=room)
            await database_sync_to_async(chat.save)()
            #broadcasting
            response = {"type": "chat.message", "message": quiz_data,"count":cache.get('user_count_'+self.group_name),"endQuiz":False}
            await self.channel_layer.group_send(self.group_name, response)
        except Exception as e:
            print('error found',e)
        #     await self.send(text_data=json.dumps({"message": "Login Required"}))

    async def chat_message(self, event):
        print("Event..", event)
        await self.send(text_data=json.dumps({"message": event["message"],"count":event["count"],"endQuiz":event["endQuiz"]}))

    async def disconnect(self, code):
        print("web socket disconnected ", code)
        print("channel_name " + self.channel_name)
        print("channel_layer ", self.channel_layer)
        try:
            online_users=online_user_count(self.group_name,'DECR')
        except:
            online_users=0
        await self.channel_layer.group_discard(self.group_name, self.channel_name)
        # sending online user count when someone left the group/room
        response = {"type": "chat.message", "message": "Someone Left","count":online_users,"endQuiz":False}
        await self.channel_layer.group_send(self.group_name, response)
